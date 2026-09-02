"""Round-trip test for the xlsx export over a tiny synthetic filled/ dir."""
from __future__ import annotations

import csv

from openpyxl import load_workbook

from psm.export_e19 import ABOUT_LINES, PROVENANCE_FILLS, _xlsx_safe, export


def _write(path, fieldnames, rows):
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)


def _tiny_filled(tmp_path):
    filled = tmp_path / "filled"
    filled.mkdir()
    icols = ["Incident Number", "Work Group", "Date of Incident",
             "Time of Incident", "Health & Safety - Risk Score"]
    _write(filled / "incidents.csv", icols, [
        {"Incident Number": "A-1", "Work Group": "Drilling",
         "Date of Incident": "2024-05-02", "Time of Incident": "16:20",
         "Health & Safety - Risk Score": "20"},
        {"Incident Number": "A-2", "Work Group": "Marine",
         "Date of Incident": "2020-01-15", "Time of Incident": "",
         "Health & Safety - Risk Score": ""},
    ])
    _write(filled / "provenance.csv", icols, [
        {"Incident Number": "src", "Work Group": "syn", "Date of Incident": "src",
         "Time of Incident": "src", "Health & Safety - Risk Score": "src"},
        {"Incident Number": "src", "Work Group": "syn", "Date of Incident": "src",
         "Time of Incident": "", "Health & Safety - Risk Score": ""},
    ])
    ccols = ["Incident Number", "Cause number", " Failed PSM Framework Element"]
    _write(filled / "causes.csv", ccols,
           [{"Incident Number": "A-1", "Cause number": "1", " Failed PSM Framework Element": "17"}])
    _write(filled / "causes_provenance.csv", ccols,
           [{"Incident Number": "src", "Cause number": "src", " Failed PSM Framework Element": "llm"}])
    _write(filled / "causes_confidence.csv",
           ["Incident Number", "Cause number", "element_confidence"],
           [{"Incident Number": "A-1", "Cause number": "1", "element_confidence": "high"}])
    return filled


def test_export_builds_five_sheets_with_provenance_shading(tmp_path):
    out = tmp_path / "out.xlsx"
    export(_tiny_filled(tmp_path), out)

    wb = load_workbook(out)
    assert wb.sheetnames == [
        "About", "Incidents", "Causes",
        "Incidents Provenance", "Causes Provenance",
    ]
    inc = wb["Incidents"]
    assert inc["A1"].value == "Incident Number"
    assert inc["B2"].value == "Drilling"
    assert inc["B2"].fill.start_color.rgb == "00" + PROVENANCE_FILLS["syn"]
    causes = wb["Causes"]
    assert causes["C2"].fill.start_color.rgb == "00" + PROVENANCE_FILLS["llm"]
    assert "not a real" in str(wb["About"]["A4"].value).lower()


def test_provenance_sheets_mirror_data_sheets_cell_for_cell(tmp_path):
    out = tmp_path / "out.xlsx"
    export(_tiny_filled(tmp_path), out)
    wb = load_workbook(out)

    inc = wb["Incidents"]
    inc_prov = wb["Incidents Provenance"]
    causes = wb["Causes"]
    causes_prov = wb["Causes Provenance"]

    # Identical headers, byte-exact, same order.
    inc_headers = [c.value for c in inc[1]]
    assert [c.value for c in inc_prov[1]] == inc_headers
    causes_headers = [c.value for c in causes[1]]
    assert [c.value for c in causes_prov[1]] == causes_headers

    # Token cell matches the fixture's provenance.csv / causes_provenance.csv.
    assert inc_prov["B2"].value == "syn"          # Incidents!B2 == "Drilling"
    assert causes_prov["C2"].value == "llm"        # Causes!C2 == 17

    # A cell whose token is empty in the CSV is genuinely None, not "".
    assert inc["D3"].value is None                 # Time of Incident, row A-2
    assert inc_prov["D3"].value is None

    assert inc_prov.freeze_panes == "B2"
    assert causes_prov.freeze_panes == "B2"


def test_header_cells_carry_provenance_comments(tmp_path):
    out = tmp_path / "out.xlsx"
    export(_tiny_filled(tmp_path), out)
    wb = load_workbook(out)

    inc = wb["Incidents"]
    b1 = inc["B1"].comment
    assert b1 is not None
    assert "syn" in b1.text and "corresponds to nothing real" in b1.text

    a1 = inc["A1"].comment
    assert a1 is not None
    assert "src" in a1.text

    for ws in (wb["Incidents"], wb["Causes"]):
        for header_cell in ws[1]:
            assert header_cell.comment is not None, (
                f"{ws.title}!{header_cell.coordinate} has no header comment"
            )


def test_about_names_provenance_sheets_and_warns_of_ai_extraction():
    text = "\n".join(ABOUT_LINES)
    assert "Incidents Provenance" in text and "Causes Provenance" in text
    assert "AI" in text
    assert "does not survive" in text
    assert "copy" in text and "export" in text


def test_dates_times_and_scores_are_typed_cells_not_text(tmp_path):
    """The SME's first move is sorting or filtering by risk score or date.
    Text-typed digits sort lexicographically ('20' < '4') and get Text
    Filters instead of Number Filters -- a silent corruption of exactly the
    fitness-for-purpose question the deliverable exists to answer."""
    import datetime

    out = tmp_path / "out.xlsx"
    export(_tiny_filled(tmp_path), out)
    wb = load_workbook(out)
    inc = wb["Incidents"]

    assert inc["C2"].value == datetime.datetime(2024, 5, 2)   # Date of Incident
    assert inc["C2"].is_date
    assert inc["D2"].value == datetime.time(16, 20)           # Time of Incident
    assert inc["E2"].value == 20                              # Risk Score, int
    assert inc["E2"].data_type == "n"

    # A blank in a typed column must be genuinely empty -- not the string ''
    # -- so AutoFilter offers (Blanks) and arithmetic skips it.
    assert inc["D3"].value is None
    assert inc["E3"].value is None

    # Typed causes columns too: element codes and cause ordinals are ints.
    assert wb["Causes"]["B2"].value == 1
    assert wb["Causes"]["C2"].value == 17


def test_unparseable_value_in_a_typed_column_fails_loudly(tmp_path):
    """A typed column with a value that does not parse must raise, never
    silently fall back to text -- a mixed-type column is worse than either."""
    import pytest

    filled = _tiny_filled(tmp_path)
    _write(filled / "incidents.csv",
           ["Incident Number", "Work Group", "Date of Incident",
            "Time of Incident", "Health & Safety - Risk Score"],
           [{"Incident Number": "A-1", "Work Group": "Drilling",
             "Date of Incident": "2024-05-02", "Time of Incident": "16:20",
             "Health & Safety - Risk Score": "not-a-number"}])
    _write(filled / "provenance.csv",
           ["Incident Number", "Work Group", "Date of Incident",
            "Time of Incident", "Health & Safety - Risk Score"],
           [{"Incident Number": "src", "Work Group": "syn", "Date of Incident": "src",
             "Time of Incident": "src", "Health & Safety - Risk Score": "src"}])
    with pytest.raises(ValueError):
        export(filled, tmp_path / "out.xlsx")


def test_first_column_and_header_row_stay_frozen(tmp_path):
    """B2 keeps both the header row and the Incident Number column in view
    while a reviewer scrolls 43 columns to the right."""
    out = tmp_path / "out.xlsx"
    export(_tiny_filled(tmp_path), out)
    wb = load_workbook(out)
    assert wb["Incidents"].freeze_panes == "B2"
    assert wb["Causes"].freeze_panes == "B2"


def test_xlsx_safe_substitutes_control_character_runs_with_a_single_space():
    # A control byte sitting at a word boundary must not be deleted outright
    # -- that would silently fuse "burn" and "injury" into "burninjury",
    # fabricating a word that isn't in the source text.
    assert _xlsx_safe("burn\x01injury") == "burn injury"

    # A run of several consecutive illegal bytes must become exactly one
    # space, not one space per byte -- otherwise a 12-byte run would pad the
    # cell with 12 spaces.
    assert _xlsx_safe("a\x01\x01\x01\x01\x01\x01\x01\x01\x01\x01\x01\x01b") == "a b"

    # Identity (not just equality) proves the sanitiser leaves clean text
    # completely alone -- it cannot quietly rewrite whitespace or characters
    # that were never illegal in the first place.
    clean = "clean cause text with normal   spacing, nothing illegal here."
    assert _xlsx_safe(clean) is clean

    # Correct behaviour: a cell whose entire content is one run of illegal
    # bytes becomes a single space, not an empty string. This follows
    # directly from "each run becomes one space" with no extra stripping --
    # an empty-string special case would be an inconsistent exception to that
    # rule, and would make a genuinely-blank source cell indistinguishable
    # from one BSEE's PDF extraction corrupted into unreadability.
    assert _xlsx_safe("\x01\x01\x01") == " "


def test_about_legend_covers_key_and_pseud_and_drops_the_stale_exception():
    text = "\n".join(ABOUT_LINES)
    assert "green" in text and "constructed" in text          # key legend line
    assert "lilac" in text and "pseudonym" in text            # pseud legend line
    # the superseded paragraph claimed Incident Number was the lone exception
    assert "Incident Number is unshaded" not in text
    # new caveat: Cause type is ordinal position, not analysis
    assert "Cause type" in text and "position" in text


def test_about_discloses_what_the_adversarial_review_found_undisclosed():
    """Each assertion below is a review finding: a true fact about the
    workbook that the About sheet did not state. The company workbooks got
    the names disclosure in final review; this workbook had not."""
    text = "\n".join(ABOUT_LINES)
    # real people appear verbatim in narrative/cause text
    assert "individuals" in text and "verbatim" in text
    # 'no colour' covers genuinely-empty cells too, not only verbatim src
    assert "no colour" in text and ("empty" in text or "blank" in text)
    # '109' is two different outcomes, kept separate in docs/findings.md
    assert "107" in text and "2 parse failures" in text
    # pseudonym tokens sit beside verbatim position titles
    assert "position" in text.lower()
    # byte-exact headers include the template's own typo -- deliberate
    assert "(sic)" in text
    # keyless rows carry UNKEYED- tokens
    assert "UNKEYED" in text
    # the triplicated classification columns are by template design
    assert "three" in text
    # control characters occur on both sheets, not only cause descriptions
    assert "both sheets" in text
