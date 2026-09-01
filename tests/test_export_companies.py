from openpyxl import load_workbook

from psm.export_companies import (ABOUT_TEMPLATE, COMPANY_LABELS,
                                  export_all)


def test_company_about_discloses_without_leaking_the_answer_key():
    # Render against every real label -- the raw ABOUT_TEMPLATE has an
    # unformatted "{label}" placeholder, so checking it directly can never
    # catch a banned token that only appears once {label} is substituted
    # (e.g. the identity statement "Coastal is a SYNTHETIC company").
    for label in COMPANY_LABELS.values():
        text = "\n".join(line.format(label=label) for line in ABOUT_TEMPLATE)
        text = text.lower()
        for needed in ("synthetic", "template", "shifted", "deliberately"):
            assert needed in text, (label, needed)
        for banned in ("pathology", "answer key", "closeout decay",
                       "no real names"):
            assert banned not in text, (label, banned)


def test_export_writes_three_company_workbooks_and_the_comparison(tmp_path):
    export_all(out_dir=tmp_path)
    for label in COMPANY_LABELS.values():
        wb = load_workbook(tmp_path / f"{label}_E19_Register.xlsx")
        assert wb.sheetnames == ["About", "Incidents", "Causes",
                                 "Recommendations", "Closeout"]
    cmp_wb = load_workbook(tmp_path / "comparison.xlsx")
    assert cmp_wb.sheetnames == ["About", "KPIs", "Plants", "Negative Controls"]
    kpi_ws = cmp_wb["KPIs"]
    assert kpi_ws.max_row == 10            # header + 9 KPIs
    assert kpi_ws.max_column == 4          # kpi name + 3 companies


def test_incident_number_column_is_shaded_key_green(tmp_path):
    export_all(out_dir=tmp_path)
    wb = load_workbook(tmp_path / "NorthStar_E19_Register.xlsx")
    ws = wb["Incidents"]
    hdr = [c.value for c in ws[1]]
    i = hdr.index("Incident Number") + 1
    fills = {ws.cell(row=r, column=i).fill.start_color.rgb
             for r in range(2, 30)}
    assert fills == {"00E2EFDA"}


def test_comparison_about_names_the_bundle_level_claim(tmp_path):
    export_all(out_dir=tmp_path)
    wb = load_workbook(tmp_path / "comparison.xlsx")
    text = "\n".join(str(r[0].value or "") for r in wb["About"].iter_rows())
    assert "bundle-level" in text
    assert "never distributed" in text
