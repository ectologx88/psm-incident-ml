"""Export the filled/ E19 layer to one xlsx for SME review.

Three sheets: About (what this is and is not), Incidents, Causes. Every cell
whose provenance token is xw/llm/syn/key/pseud carries a fill colour so a
reviewer can see at a glance which values are real, mapped, model-assigned,
synthetic, a constructed identifier, or a pseudonym (legend on the About
sheet). The workbook is a DELIVERABLE, not a dataset of record —
deliverables/ is gitignored; the committed record is
data/processed/e19/filled/*.csv plus the parallel provenance files.

Run:  uv run python -m psm.export_e19
"""
from __future__ import annotations

import csv
import datetime
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill

from psm.fill import FILLED
from psm.provenance import FILL_COLORS as PROVENANCE_FILLS
from psm.provenance import UNSHADED

REPO = Path(__file__).resolve().parents[2]
DEFAULT_OUT = REPO / "deliverables" / "e19_filled.xlsx"

ABOUT_LINES = [
    "E19 Investigation Register - filled demonstration copy",
    "",
    "Built from public US BSEE offshore incident reports, projected into the",
    "Energy Institute PSM Framework Element 19 register shape. This is not a real",
    "filled E19 worksheet: it demonstrates what an auto-populated register looks",
    "like. Cell colours state where every value came from:",
    "",
    "  no colour  - verbatim from a BSEE source document, or genuinely empty",
    "               (BSEE recorded nothing and this repo declined to invent",
    "               it -- the blank is the deliverable)",
    "  blue       - deterministic crosswalk from a BSEE category (an opinion,",
    "               recorded in schema/crosswalk.yaml)",
    "  amber      - assigned by a language model (3-pass self-consistency,",
    "               Claude Haiku 4.5) - never treated as ground truth",
    "  grey       - synthetic: deterministic hash-generated filler for fields",
    "               BSEE does not publish (names are SYN- tokens, dates are",
    "               offsets, picklist values are invented). Corresponds to",
    "               nothing real.",
    "  green      - constructed identifier: BSEE publishes no incident id,",
    "               so this repo builds the key (area-block-date-time, some",
    "               with content-hash parts; rows where those parts are",
    "               unavailable carry an UNKEYED-<hash> token). Consistent,",
    "               but corresponds to no source field.",
    "  lilac      - salted pseudonym of a real name (INV-/SUP- tokens).",
    "               Same person, same token, corpus-wide. De-amplification",
    "               of public documents, not fabrication.",
    "",
    "Model-assigned labels are unvalidated. On the 524 statements where the",
    "crosswalk also holds an opinion, the model agreed on 25.4% (133) and",
    "produced no label on 109 (107 abstentions, 2 parse failures); counting",
    "only the 415 where both produced a label, agreement is 32.0%. The corpus",
    "also skews heavily to one category. Treat every amber/grey cell as a",
    "proposal to evaluate, not a finding. Full provenance:",
    "data/processed/e19/filled/ in the psm-incident-ml repository.",
    "",
    "The Cause type column reflects ordinal position in the source list",
    "(cause #1 is always 'Immediate'), not causal analysis. Do not read",
    "root-cause depth from it.",
    "",
    "Disclosures:",
    "- Structured name fields are SYN-/INV-/SUP- tokens; narrative and",
    "  cause text is verbatim public BSEE report text and names the real",
    "  operators, vessels, facilities and -- occasionally -- the",
    "  individuals involved, including injured or implicated parties.",
    "- A pseudonym (lilac) can sit in the same row as a verbatim position",
    "  title, and the pairing narrows who a token could refer to. The",
    "  tokens de-amplify public documents; they are not anonymisation.",
    "- Column headers reproduce the E19 template byte-exact, including its",
    "  own irregularities: 'Incident Classificatioin' (sic) and stray",
    "  spaces are the template's, not transcription errors made here.",
    "- Incident Classification appears three times by template design",
    "  (columns C, AC and AE); all three carry the same value on every row.",
    "- Dates, times, risk scores, likelihoods and cause/element codes are",
    "  written as typed date/time/number cells so sorting and Number",
    "  Filters behave; the committed CSVs remain the textual record.",
    "- A few narrative and cause-description cells (on both sheets) contain",
    "  control characters left by PDF extraction that xlsx cannot store;",
    "  each run is rendered here as a single space. The committed CSVs in",
    "  data/processed/e19/filled/ keep the original bytes.",
]

# The About legend's coloured lines get a swatch cell in column B, filled
# with the actual PatternFill the data sheets use -- so the legend cannot
# drift from the real colours, and a reader who cannot distinguish the hues
# by name can match swatch to cell directly.
_LEGEND_SWATCHES = {
    "  blue": "xw",
    "  amber": "llm",
    "  grey": "syn",
    "  green": "key",
    "  lilac": "pseud",
}


def _rows(path: Path) -> tuple[list[str], list[dict]]:
    with path.open(encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        return list(reader.fieldnames or []), list(reader)


# One or more consecutive illegal characters, collapsed to a single space by
# _xlsx_safe rather than one space per byte (openpyxl's own
# ILLEGAL_CHARACTERS_RE has no quantifier, so it would otherwise match -- and
# get substituted -- one character at a time).
_ILLEGAL_RUN_RE = re.compile(f"(?:{ILLEGAL_CHARACTERS_RE.pattern})+")


def _xlsx_safe(value: str) -> str:
    """Substitute each run of characters OOXML text cells cannot hold at all
    (e.g. stray \\x01 bytes from PDF extraction — see docs/findings.md,
    "source data is dirty and stays dirty") with a single space, rather than
    deleting them. In this corpus a control byte sits at a word boundary
    (e.g. "burn\\x01injury\\x01to"), so deleting it would silently fabricate a
    word ("burninjuryto") that never appeared in the source; a space keeps the
    text truthful. A run of several consecutive control bytes collapses to
    one space, not one space per byte. No other whitespace is touched: no
    stripping, no collapsing of spaces that were already in the source text.

    This is an xlsx-format compatibility step, not a data edit: the CSVs
    under data/processed/e19/filled/ remain the untouched record with the
    original bytes; only this deliverable's rendering of them is sanitised.
    """
    return _ILLEGAL_RUN_RE.sub(" ", value)


# Columns written as typed cells rather than text. Text-typed digits sort
# lexicographically ('20' < '4') and draw Text Filters instead of Number
# Filters -- silently wrong for exactly the sort-by-risk-score move a
# reviewer makes first. The lists are explicit rather than inferred from the
# data so a schema change cannot silently retype a column; a value that does
# not parse raises instead of falling back to text (a mixed-type column is
# worse than either).
INT_COLS = frozenset({
    "Health & Safety - Risk Score",
    "Health & Safety - Likelihood",
    "Environment & Reputation - Risk Score",
    "Environment & Reputation - Likelihood",
    "Financial Cost & Business Interruption - Risk Score",
    "Financial Cost & Business Interruption - Likelihood",
    "Cause number",
    " Failed PSM Framework Element",
})
DATE_COLS = frozenset({
    "Date of Incident", "Date of Report", "Approval Date", "Close out Date",
})
TIME_COLS = frozenset({"Time of Incident"})

# Narrative columns whose content routinely exceeds one screen-width; they
# get wrap_text so a row's prose is readable in place instead of overflowing
# into (or hiding behind) its neighbours.
_WRAP_MIN_CHARS = 80


def _typed(col: str, raw: str) -> object:
    """Parse ``raw`` for a typed column; text columns pass through _xlsx_safe.

    A blank in a typed column becomes None -- a genuinely empty cell -- so
    AutoFilter offers (Blanks) and arithmetic skips it, rather than an
    empty-string text cell that reads as a value.
    """
    if col in INT_COLS:
        return int(raw) if raw.strip() else None
    if col in DATE_COLS:
        return datetime.date.fromisoformat(raw) if raw.strip() else None
    if col in TIME_COLS:
        return datetime.time.fromisoformat(raw) if raw.strip() else None
    return _xlsx_safe(raw)


def _write_sheet(ws, cols: list[str], rows: list[dict], prov: list[dict]) -> int:
    """Returns the number of cells whose value was sanitised by _xlsx_safe."""
    assert len(rows) == len(prov), "value/provenance row count mismatch"
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    typed = INT_COLS | DATE_COLS | TIME_COLS
    wrap_cols = {
        c for c in cols
        if c not in typed and any(len(r[c]) > _WRAP_MIN_CHARS for r in rows)
    }
    wrap = Alignment(wrap_text=True, vertical="top")
    sanitised = 0
    for row, prow in zip(rows, prov):
        values = []
        for c in cols:
            raw = row[c]
            try:
                value = _typed(c, raw)
            except ValueError as exc:
                raise ValueError(f"{c!r}: cannot type value {raw!r}") from exc
            if isinstance(value, str) and value != raw:
                sanitised += 1
            values.append(value)
        ws.append(values)
        for j, c in enumerate(cols, start=1):
            token = prow.get(c, "")
            if token not in PROVENANCE_FILLS and token not in UNSHADED:
                raise ValueError(
                    f"{c!r}: unknown provenance token {token!r} -- "
                    "not in PROVENANCE_FILLS and not in UNSHADED"
                )
            cell = ws.cell(row=ws.max_row, column=j)
            if token in PROVENANCE_FILLS:
                cell.fill = PatternFill("solid", start_color=PROVENANCE_FILLS[token])
            if c in DATE_COLS:
                cell.number_format = "yyyy-mm-dd"
            elif c in TIME_COLS:
                cell.number_format = "hh:mm"
            elif c in wrap_cols:
                cell.alignment = wrap
    # Post-write assertion: every populated cell of a typed column really is
    # typed. If _typed ever regresses to passing strings through, this fails
    # the export rather than shipping a silently text-typed workbook.
    for j, c in enumerate(cols, start=1):
        if c not in typed:
            continue
        for cell in next(ws.iter_cols(min_col=j, max_col=j, min_row=2)):
            assert cell.value is None or cell.data_type in ("n", "d"), (
                f"{c!r}: cell {cell.coordinate} is {cell.data_type!r}, not typed"
            )
    # B2, not A2: keep the header row AND the Incident Number column in view
    # while a reviewer scrolls the 40-odd columns to the right.
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    for j, c in enumerate(cols, start=1):
        width = min(max(len(c) + 2, 12), 60)
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = width
    return sanitised


def export(filled_dir: Path, out_path: Path) -> int:
    """Builds the workbook and returns the number of cells sanitised by
    _xlsx_safe, so callers can observe (and log/print) how many cells were
    altered instead of that fact being visible only via a print statement."""
    wb = Workbook()
    about = wb.active
    about.title = "About"
    for i, line in enumerate(ABOUT_LINES, start=1):
        about.append([line])
        for prefix, token in _LEGEND_SWATCHES.items():
            if line.startswith(prefix + " "):
                about.cell(row=i, column=2).fill = PatternFill(
                    "solid", start_color=PROVENANCE_FILLS[token]
                )
    about.column_dimensions["A"].width = 90
    about.column_dimensions["B"].width = 4

    icols, irows = _rows(filled_dir / "incidents.csv")
    _, iprov = _rows(filled_dir / "provenance.csv")
    sanitised = _write_sheet(wb.create_sheet("Incidents"), icols, irows, iprov)

    ccols, crows = _rows(filled_dir / "causes.csv")
    _, cprov = _rows(filled_dir / "causes_provenance.csv")
    sanitised += _write_sheet(wb.create_sheet("Causes"), ccols, crows, cprov)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return sanitised


def main() -> int:
    sanitised = export(FILLED, DEFAULT_OUT)
    print(f"wrote {DEFAULT_OUT}")
    print(f"sanitised {sanitised} cells containing control characters unrepresentable in xlsx")
    print("deliverable only - never commit; the record is data/processed/e19/filled/")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
