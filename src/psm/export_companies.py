"""Export the three synthetic-company registers to reviewer workbooks, plus
the INTERNAL comparison workbook (the answer key -- never distributed with
the company workbooks).

Run:  uv run python -m psm.export_companies
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from psm.export_e19 import _write_sheet
from psm.kpi import compute_kpis, load_company
from psm.scenario import COMPANY_ORDER, OUT_ROOT

REPO = Path(__file__).resolve().parents[2]
OUT_DIR = REPO / "deliverables" / "companies"

COMPANY_LABELS = {"northstar": "NorthStar", "meridian": "Meridian",
                  "coastal": "Coastal"}

_SHEETS = (("Incidents", "incidents", "provenance.csv"),
           ("Causes", "causes", "causes_provenance.csv"),
           ("Recommendations", "recommendations",
            "recommendations_provenance.csv"),
           ("Closeout", "closeout", "closeout_provenance.csv"))

ABOUT_TEMPLATE = [
    "{label} Offshore -- E19 Investigation Register (synthetic demonstration)",
    "",
    "{label} is a SYNTHETIC company. This register was generated from public",
    "US BSEE offshore incident narratives, rebased into 2021-2025 at a",
    "large-operator scale (~30 incidents/yr), with process health",
    "deliberately varied between the companies in this evaluation set so an",
    "incident-management evaluator can be tested against known conditions.",
    "Which conditions were varied, and where, is documented separately and",
    "intentionally not stated here.",
    "",
    "Cell colours state provenance (same scheme as the source project):",
    "  no colour  - verbatim text from a public BSEE incident report",
    "  grey       - synthetic: deterministic generated value (dates, names,",
    "               picklists, recommendation text). Corresponds to nothing real.",
    "  green      - constructed identifier (this register's own keys)",
    "",
    "Disclosures:",
    "- Dates inside narratives were uniformly shifted with each incident's",
    "  rebased timeline, so prose and register agree; era-distant dates in",
    "  the source text (including OCR debris) were left as found.",
    "- Recommendation text comes from a fixed template registry adapted from",
    "  real operator-voice recommendations; repetition across incidents is",
    "  intentional corporate boilerplate.",
    "- People are SYN- tokens. No real names appear in this register.",
]

_COMPARISON_ABOUT = [
    "Scenario comparison workbook -- INTERNAL VALIDATION ARTIFACT",
    "",
    "This is the answer key for the synthetic-company registers: it names",
    "the planted process pathologies, the measured KPI values, and the",
    "negative-control checks. It is never distributed alongside the company",
    "workbooks.",
    "",
    "Attribution honesty: Coastal's pathologies co-move by construction, so",
    "results for Coastal are claimed at bundle-level detection only --",
    "per-pathology attribution is out of scope (parked in the spec).",
]


def _about_sheet(wb: Workbook, lines: list[str]) -> None:
    ws = wb.active
    ws.title = "About"
    for line in lines:
        ws.append([line])
    ws.column_dimensions["A"].width = 90


def _company_workbook(company: str, out_dir: Path) -> Path:
    import csv
    label = COMPANY_LABELS[company]
    src = OUT_ROOT / company
    wb = Workbook()
    _about_sheet(wb, [ln.format(label=label) for ln in ABOUT_TEMPLATE])
    for sheet, table, prov_file in _SHEETS:
        with (src / f"{table}.csv").open(encoding="utf-8", newline="") as fh:
            reader = csv.DictReader(fh)
            cols, rows = list(reader.fieldnames or []), list(reader)
        with (src / prov_file).open(encoding="utf-8", newline="") as fh:
            prov = list(csv.DictReader(fh))
        _write_sheet(wb.create_sheet(sheet), cols, rows, prov)
    path = out_dir / f"{label}_E19_Register.xlsx"
    wb.save(path)
    return path


def _comparison_workbook(out_dir: Path) -> Path:
    kpis = {c: compute_kpis(load_company(OUT_ROOT / c)) for c in COMPANY_ORDER}
    manifests = {c: json.loads((OUT_ROOT / c / "manifest.json")
                               .read_text(encoding="utf-8"))
                 for c in COMPANY_ORDER}
    wb = Workbook()
    _about_sheet(wb, _COMPARISON_ABOUT)

    ws = wb.create_sheet("KPIs")
    ws.append(["KPI"] + [COMPANY_LABELS[c] for c in COMPANY_ORDER])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for kpi in sorted(kpis[COMPANY_ORDER[0]]):
        ws.append([kpi] + [round(float(kpis[c][kpi]), 4)
                           for c in COMPANY_ORDER])

    ws = wb.create_sheet("Plants")
    ws.append(["Company", "Pathology", "KPI", "Expected", "Measured"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for c in COMPANY_ORDER:
        for p in manifests[c]["plants"]:
            ws.append([COMPANY_LABELS[c], p["pathology"], p["kpi"],
                       json.dumps(p["expected"]),
                       round(float(kpis[c][p["kpi"]]), 4)])

    ws = wb.create_sheet("Negative Controls")
    ws.append(["Company", "Control"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for c in COMPANY_ORDER:
        for ctl in manifests[c]["negative_controls"]:
            ws.append([COMPANY_LABELS[c], ctl])

    path = out_dir / "comparison.xlsx"
    wb.save(path)
    return path


def export_all(out_dir: Path = OUT_DIR) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    written = [_company_workbook(c, out_dir) for c in COMPANY_ORDER]
    written.append(_comparison_workbook(out_dir))
    return written


def main() -> int:
    for path in export_all():
        print(f"wrote {path}")
    print("deliverables only - never commit; the record is data/companies/")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
