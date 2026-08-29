"""Extract the E19 target schema from the source workbook, mechanically.

The E19 workbook is a workplace document and is NOT committed (see CLAUDE.md).
This module reads it read-only and emits ``schema/e19_labels.yaml``, which is
committed: field labels and picklist vocabularies only, no formulas, no scoring
logic, no rollup structure.

Why this exists rather than a hand-written schema: every field-name mismatch
found in review so far came from a human transcribing labels. Labels are read
from the workbook and never retyped, including their irregularities --
``Incident Classificatioin`` (sic), ``Unmittigated`` (sic), ``Human Factors``
with a double space, `` Failed PSM Framework Element`` with a leading space.
Those are the real column names; "correcting" them would break the exactness
guarantee the projection layer exists to provide.

Usage::

    uv run python -m psm.e19_schema --workbook /path/to/"E19 Investigation Report - Rev2.xlsx"

The workbook path is deliberately a required argument with no default, so the
file is never assumed to live inside the repo.
"""

from __future__ import annotations

import argparse
import hashlib
import sys
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import yaml

# Sheets read. Both are read-only; the workbook is never written to.
SHEET_FIELDS = "Database Fields"
SHEET_PICKLISTS = "Dropdown Picklist Data"

# Columns on the Database Fields sheet that carry labels. Derived by scanning,
# not hardcoded row numbers -- see _blocks_in_column.
LABEL_COLUMNS = ("E", "I", "M", "Q")

MAX_ROW = 200
MAX_PICKLIST_COL = 60
MAX_PICKLIST_ROW = 80


@dataclass
class Block:
    """A labelled group of fields: one header row, then a run of field rows."""

    header: str
    header_cell: str
    fields: list[dict] = field(default_factory=list)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def _runs(rows: list[int]) -> list[list[int]]:
    """Split a sorted row list into runs of contiguous rows."""
    out: list[list[int]] = []
    for r in rows:
        if out and r == out[-1][-1] + 1:
            out[-1].append(r)
        else:
            out.append([r])
    return out


def _blocks_in_column(ws, col: str) -> list[Block]:
    """Find header/field blocks in one column.

    The sheet's own layout convention is: a header cell, a blank row, then a
    contiguous run of field labels. So a run of length 1 is a header and the
    run that follows it holds that header's fields. Detecting this rather than
    hardcoding row numbers means the extractor survives the author inserting a
    row.
    """
    populated = [
        r
        for r in range(1, MAX_ROW + 1)
        if (v := ws[f"{col}{r}"].value) is not None and str(v).strip() != ""
    ]
    blocks: list[Block] = []
    runs = _runs(populated)
    i = 0
    while i < len(runs):
        run = runs[i]
        if len(run) == 1 and i + 1 < len(runs):
            header_row = run[0]
            blocks.append(
                Block(
                    header=str(ws[f"{col}{header_row}"].value),
                    header_cell=f"{col}{header_row}",
                    fields=[
                        {
                            "label": str(ws[f"{col}{r}"].value),
                            "cell": f"{col}{r}",
                        }
                        for r in runs[i + 1]
                    ],
                )
            )
            i += 2
        else:
            # A run with no following run, or a multi-row run with no header:
            # record it so nothing is silently dropped.
            blocks.append(
                Block(
                    header=str(ws[f"{col}{run[0]}"].value),
                    header_cell=f"{col}{run[0]}",
                    fields=[
                        {"label": str(ws[f"{col}{r}"].value), "cell": f"{col}{r}"}
                        for r in run[1:]
                    ],
                )
            )
            i += 1
    return blocks


def read_fields(wb) -> list[dict]:
    ws = wb[SHEET_FIELDS]
    out: list[dict] = []
    for col in LABEL_COLUMNS:
        for blk in _blocks_in_column(ws, col):
            out.append(
                {
                    "group": blk.header,
                    "group_cell": blk.header_cell,
                    "fields": blk.fields,
                }
            )
    return out


# The investigation-form vocabularies sit on a common header row. Columns whose
# content starts below this are bare continuation lists with no name of their
# own -- treating their first cell as a vocabulary name invents a label that is
# really a value. Observed: X45 and AI6 are unnumbered/numbered PSM element
# lists; BF7/BH7 are rating scales.
VOCAB_HEADER_MAX_ROW = 5


def _is_numberish(v) -> bool:
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).strip())
    except (TypeError, ValueError):
        return False
    return True


def _split_sequence_restarts(pairs: list[tuple[int, object]]) -> list[list[tuple[int, object]]]:
    """Split one column into runs, breaking where a numeric series restarts.

    The 'Risk Score' column stacks 1..25 immediately followed by 1..20 (the PSM
    element numbers) with no blank row between them. Read naively that is a
    single 45-value vocabulary, which is wrong in a way nothing downstream would
    catch.
    """
    runs: list[list[tuple[int, object]]] = [[]]
    prev: float | None = None
    for r, v in pairs:
        if _is_numberish(v):
            cur = float(str(v).strip())
            if prev is not None and cur <= prev:
                runs.append([])
                prev = None
            if runs[-1]:
                prev = cur
            else:
                prev = cur
        else:
            prev = None
        runs[-1].append((r, v))
    return [run for run in runs if run]


def read_picklists(wb) -> list[dict]:
    """Read vocabularies verbatim -- including 'Permenant', 'Electricution',
    'distrating', 'Availiability', 'Spil to Land'.

    Emits ``header_confident`` rather than guessing. A column is only treated as
    a named vocabulary when its first cell sits on the sheet's header band and
    is not itself numeric; otherwise ``name`` is null and the consumer must
    identify the list rather than trust an invented name.
    """
    ws = wb[SHEET_PICKLISTS]
    out: list[dict] = []
    for c in range(1, MAX_PICKLIST_COL + 1):
        letter = openpyxl.utils.get_column_letter(c)
        pairs = [
            (r, ws.cell(row=r, column=c).value) for r in range(1, MAX_PICKLIST_ROW + 1)
        ]
        pairs = [(r, v) for r, v in pairs if v is not None and str(v).strip() != ""]
        if len(pairs) < 2:
            continue

        for run in _split_sequence_restarts(pairs):
            if len(run) < 2:
                continue
            first_row, first_val = run[0]
            confident = first_row <= VOCAB_HEADER_MAX_ROW and not _is_numberish(first_val)
            name, values = (str(first_val), run[1:]) if confident else (None, run)
            out.append(
                {
                    "name": name,
                    "header_confident": confident,
                    "cell": f"{letter}{first_row}",
                    # Rows are stored, not inferred: several columns have blank
                    # rows inside them, so a contiguous range would be wrong.
                    "value_rows": [r for r, _ in values],
                    "values": [str(v) for _, v in values],
                }
            )
    return out


def build(workbook: Path) -> dict:
    digest_before = sha256(workbook)
    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=False)
    doc = {
        "_generated_by": "src/psm/e19_schema.py",
        "_source_workbook": workbook.name,
        "_source_sha256": digest_before,
        "_note": (
            "Labels are byte-exact from the workbook, including its own typos and "
            "irregular whitespace. Do not normalise them: these are the column "
            "names the E19 projection must emit. Regenerate with "
            "`uv run python -m psm.e19_schema --workbook <path>` rather than editing."
        ),
        "groups": read_fields(wb),
        "vocabularies": read_picklists(wb),
    }
    wb.close()
    if sha256(workbook) != digest_before:
        raise RuntimeError("workbook changed during read -- aborting")
    return doc


def verify(doc: dict, workbook: Path) -> list[str]:
    """Round-trip: every emitted label must still match its source cell exactly."""
    problems: list[str] = []
    wb = openpyxl.load_workbook(workbook, data_only=True)
    ws = wb[SHEET_FIELDS]
    for grp in doc["groups"]:
        if str(ws[grp["group_cell"]].value) != grp["group"]:
            problems.append(f"group mismatch at {grp['group_cell']}")
        for f in grp["fields"]:
            if str(ws[f["cell"]].value) != f["label"]:
                problems.append(f"label mismatch at {f['cell']}: {f['label']!r}")
    wsp = wb[SHEET_PICKLISTS]
    for vocab in doc["vocabularies"]:
        if vocab["header_confident"] and str(wsp[vocab["cell"]].value) != vocab["name"]:
            problems.append(f"vocabulary name mismatch at {vocab['cell']}")
        col = "".join(ch for ch in vocab["cell"] if ch.isalpha())
        for row, val in zip(vocab["value_rows"], vocab["values"], strict=True):
            if str(wsp[f"{col}{row}"].value) != val:
                problems.append(f"vocabulary value mismatch at {col}{row}: {val!r}")
                break
    wb.close()
    return problems


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", required=True, type=Path)
    ap.add_argument("--out", type=Path, default=Path("schema/e19_labels.yaml"))
    args = ap.parse_args(argv)

    if not args.workbook.exists():
        print(f"workbook not found: {args.workbook}", file=sys.stderr)
        return 2

    doc = build(args.workbook)
    problems = verify(doc, args.workbook)
    if problems:
        for p in problems:
            print(f"ROUND-TRIP FAILURE: {p}", file=sys.stderr)
        return 1

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", encoding="utf-8") as fh:
        yaml.safe_dump(doc, fh, sort_keys=False, allow_unicode=True, width=100)

    n_fields = sum(len(g["fields"]) for g in doc["groups"])
    print(f"{args.out}: {len(doc['groups'])} groups, {n_fields} fields, "
          f"{len(doc['vocabularies'])} vocabularies")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
